#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Apr  22 16:31:12 2022

@author: Aidan Pavao for Massachusetts Host-Microbiome Center
 - Run dynamic flux balance analysis (dFBA) with NMR constraints
 - Use python 3.8+ for best results
 - See /nmr-cdiff/venv/requirements.txt for dependencies

Copyright 2022 Massachusetts Host-Microbiome Center

   Licensed under the Apache License, Version 2.0 (the "License");
   you may not use this file except in compliance with the License.
   You may obtain a copy of the License at

       http://www.apache.org/licenses/LICENSE-2.0

   Unless required by applicable law or agreed to in writing, software
   distributed under the License is distributed on an "AS IS" BASIS,
   WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
   See the License for the specific language governing permissions and
   limitations under the License.

"""
import argparse
import datetime
import json
import os

import cobra as cb
from matplotlib import pyplot as plt
import matplotlib.colors
import numpy as np
import openpyxl as xl
import pandas as pd

from curveshapes import Metabolite
from standard_curves import compute_standard_curves
from synchronize import synchronizers
from trajectories import fit_trajectories
from get_color import get_cmap

SCDIR = os.path.dirname(os.path.abspath(__file__))   # location of script
BOLD = xl.styles.Font(bold=True)
MODEL_PATH = f'{SCDIR}/../../data/icdf843.json'
METHODS = {
    'fba': lambda m: m.optimize(),
    'fva': cb.flux_analysis.flux_variability_analysis,
    'loopless': cb.flux_analysis.loopless.loopless_solution,
    'pfba': cb.flux_analysis.pfba,
}

class MetaboliteCollection:
    def __init__(self):
        self.name_map = {}
        self.id_map = {}

    def new(self, met_id, met_name, substrate_name, scale):
        if (met_name not in self.name_map) and (met_id not in self.id_map):
            met = Metabolite(met_id, met_name)
            self.name_map[met.name] = met
            self.id_map[met.id] = met
        else:
            met = self.get_by_id(met_id)
        met.set_substrate_scale(substrate_name, scale)

    def get(self, met_id):
        return self.get_by_id(met_id)

    def get_by_id(self, met_id):
        if met_id in self.id_map:
            return self.id_map[met_id]
        else:
            raise ValueError(f"Collection does not contain metabolite {met_id}.")
        
    def has_id(self, met_id):
        return met_id in self.id_map
    
    def remove_met(self, met_id):
        if self.has_id(met_id):
            self.pop(met_id)
 
    def get_by_name(self, name):
        if name in self.name_map:
            return self.name_map[name]
        else:
            raise ValueError(f"Collection does not contain metabolite {name}.")
        
    def get_values(self):
        return self.id_map.values()
    
    def get_ids(self):
        return self.id_map.keys()

    def get_items(self):
        return self.id_map.items()
    
    def pop(self, mid):
        item = self.id_map.pop(mid)
        self.name_map.pop(item.name)
        return item

class Substrate():
    def __init__(self, name, label, model_id, concentration, curve, experiments, products,
                 product_ids, product_curves, from_standards, plot_ticks=None):
        self.name = name
        self.label = label
        self.model_id = model_id
        self.experiments = experiments
        self.cx = concentration
        self.curve = curve
        self.products = products
        self.product_ids = product_ids
        self.product_curves = product_curves
        self.from_standards = from_standards
        self.plot_ticks = plot_ticks

    @classmethod
    def from_json(cls, jobj, metabolite_collection: MetaboliteCollection):
        required_fields = ["name", "label", "model_id", "concentration", "curve", 
                           "experiments", "products"]
        positional_args = []
        try:
            substrate_name = jobj["name"]
            for field in required_fields:
                positional_args.append(jobj.pop(field))
        except KeyError:
            print(f"Field {field} missing for substrate {substrate_name}.")
            raise
        products = positional_args.pop(6)
        product_ids = {k: v["model_id"] for k, v in products.items()}
        product_curves = {k: v["curve"] for k, v in products.items()}
        from_standards = ("standards" in jobj)
        if from_standards:
            fpath = parse_filepath(jobj.pop("standards"))
            product_scale = compute_standard_curves(filepath=fpath, substrate=substrate_name)
        else:
            try:
                product_scale = {k: v['scale'] for k, v in products.items()}
            except KeyError:
                print("Expected field \"scale\" for all products of metabolite " \
                      + f"{substrate_name} where path to standards file was not proveded, " \
                      + "but at least one \"scale\" field was missing.")
                print("Please provide either scale fields or a standards field.")
                raise
        metabolite_collection.new(positional_args[2], substrate_name, substrate_name, 1.)
        for metname, metdata in products.items():
            metabolite_collection.new(metdata["model_id"], metname, substrate_name, product_scale[metname])
        return cls(*positional_args, product_scale, product_ids, product_curves, from_standards, **jobj)
    
    def metmap(self):
        met_dict = {self.name: self.model_id}
        met_dict.update({k: v for k, v in self.product_ids.items()})
        return met_dict

def set_bounds(model, rid, lower=0., upper=1000., update=True):
    """Set upper and lower flux bounds for a reaction.

    Parameters:
    model -- COBRA model
    rid -- the ID of the reaction to bound
    lower -- lower bound value to set (default: 0)
    upper -- upper bound value to set (default: 1000)
    update -- whether to set the bounds (default: True)
    """
    if update:
        rxn = model.reactions.get_by_id(rid)
        rxn.bounds = (lower, upper)

def reverse_flux(flux, lb, ub):
    """Reverse flux direction and swap upper/lower bounds."""
    return -1*flux, -1*ub, -1*lb

def update_uptake_bounds(model, t, met, params, update=True):
    """Calculate and set exchange reaction bounds at timepoint t.

    Parameters:
    model -- COBRA model
    t -- the individual timepoint (int or float)
    met -- the metabolite for which the exchange bounds are to be changed
    popts -- optimal logistic coefficients
    perrs -- logistic coefficient SEMs
    update -- whether to set the bounds (default: True)

    Returns optimal values and 95% confidence interval of exchange fluxes and
    estimated concentrations at the timepoint. If update==True, also sets the
    exchange reaction bounds to the 95% confidence interval limits.
    """
    # Calculate logistic solutions with flexible number of parameters
    signal, serr, exch, exerr = params.get_sol(t)
    exch_l, exch_u, signal_l, signal_u = params.get_bounds(t)
    # Reverse direction and set bounds for secretion reactions
    if met in ['glc', 'proL', 'leuL', 'valL', 'ileL', 'thrL']:
        rid = 'Ex_' + met
        exch, exch_l, exch_u = reverse_flux(exch, exch_l, exch_u)
        if met in ['leuL']:
            set_bounds(model, rid, update=update) # leave leucine unbounded
        else:
            set_bounds(model, rid, lower=exch_l, upper=exch_u, update=update)
    # Ignore data from 1H spectra
    elif met == 'acoa':
        pass
    # Update bounds for products
    else:
        rid = 'Sec_' + met
        lb = max(exch_l, 0) # do not allow reverse flux
        ub = max(exch_u, 0)
        set_bounds(model, rid, lower=lb, upper=ub, update=update)
    # Update Wood-Ljungdahl Pathway bounds with butyrate
    if met == 'but':
        wlp_l, wlp_u, _, _ = params.get_bounds(t, substrate="Glucose")
        set_bounds(model, 'ID_326', lower=0, upper=max(wlp_u, 0), update=update)
    # Allow natural abundance acetate from cysteine
    if met == 'ac':
        cys_l, cys_u, _, _ = params.get_bounds(t, substrate="Acetate13C")
        set_bounds(model, 'Ex_cysL', lower=max(cys_l, 0), upper=max(cys_u, 0), update=update)
    return exch, exch_l, exch_u, signal, signal_l, signal_u

def areaplot(df, dl, du, t_max=48, ylabel='flux (mol/gDW/h)'):
    """Plot lines with shaded confidence interval.

    Parameters:
    df -- dataframe of vectors corresponding to optimal values
    dl -- dataframe of vectors corresponding to lower bounds
    du -- dataframe of vectors corresponding to upper bounds
    t_max -- maximum value of x-axis
    ylabel -- y-axis label
    """
    f = plt.figure(figsize=(14, 10), )#fontsize=30,
    ax = plt.axes(
        xticks=(range(0, t_max+1, 12)),
        xlim=(0, t_max),
        ylim=(0, 1.05*max(ele.to_numpy().max() for ele in [df, dl, du])),
    )
    for met, ser in df.items():
        ser.index.name = 'index'
        p = ax.plot(ser.index.to_numpy(), ser.to_numpy(), '-', label=met, lw=5)
        color = matplotlib.colors.to_rgb(p[0].get_color())
        color = color + (0.2,)
        ax.fill_between(
            dl.index.to_numpy(), 
            dl[met].to_numpy(), 
            du[met].to_numpy(), 
            color=color
        )
    ax.set_xlabel('time (h)', fontsize=30, fontweight='bold')
    ax.set_ylabel(ylabel, fontsize=30, fontweight='bold')
    plt.xticks(ax.get_xticks(), weight='bold')
    plt.yticks(ax.get_yticks(), weight='bold')
    plt.legend()
    plt.show()

def niceplot(df, t_max=48, ylabel='flux (mol/gDW/h)'):
    """Plot lines.

    Parameters:
    df -- dataframe of data to plot
    t_max -- maximum value of x-axis
    ylabel -- y-axis label
    """
    ax = df.plot(
        figsize=(14, 10),
        xticks=(range(0, t_max+1, 12)),
        xlim=(0, t_max),
        ylim=(0, None),
        fontsize=30,
        lw=5,
    )
    ax.set_xlabel('time (h)', fontsize=30, fontweight='bold')
    ax.set_ylabel(ylabel, fontsize=30, fontweight='bold')
    plt.xticks(ax.get_xticks(), weight='bold')
    plt.yticks(ax.get_yticks(), weight='bold')
    # plt.title("dFBA (10/h), ATP, WLP > 8h")
    plt.legend()
    plt.show()

def areaplot2(t, substrates, params: MetaboliteCollection):
    """Plot lines with shaded confidence interval.

    Parameters:
    df -- dataframe of vectors corresponding to optimal values
    dl -- dataframe of vectors corresponding to lower bounds
    du -- dataframe of vectors corresponding to upper bounds
    t_max -- maximum value of x-axis
    ylabel -- y-axis label
    """
    cmap = get_cmap()
    figmap = {}
    for substrate in substrates:
        figmap[substrate.name] = plt.figure(figsize=(2.5, 2), constrained_layout=True)
        ax = figmap[substrate.name].add_subplot(111)
        afont = {'fontname': 'Arial', 'size': 7}
        ax.set_xlabel("Time (h)", **afont)
        ax.set_ylabel("Estimated Concentration (mM)", **afont)
        ax.set_xlim((0, 36))
        if substrate.plot_ticks is not None:
            start = substrate.plot_ticks["start"]
            stop = substrate.plot_ticks["stop"]
            major = substrate.plot_ticks["major_step"]
            minor = substrate.plot_ticks["minor_step"]
            ax.set_yticks(np.arange(start, stop+major, major))
            ax.set_yticks(np.arange(start, stop+minor, minor), minor=True)
        ax.set_xticks([0, 12, 24, 36])
        ax.set_xticks(list(range(37)), minor=True)
        ax.set_xticklabels(ax.get_xticks(), **afont)
        ax.set_yticklabels(ax.get_yticks(), **afont)
        ax.xaxis.set_tick_params(width=0.5)
        ax.yaxis.set_tick_params(width=0.5)
        plt.setp(ax.spines.values(), linewidth=0.5)
    for cpdset in params.get_values():
        for substrate_name, curveset in cpdset.logistic_sets.items():
            f = figmap[substrate_name]
            ax = f.get_axes()[0]
            ser = curveset.get_sol(t)[0]
            _, _, lb, ub = curveset.get_bounds(t)
            color = cmap[cpdset.name]
            ax.plot(t, ser, '-', label=cpdset.name, lw=2, c=color)
            color = color + (0.2,)
            ax.fill_between(t, lb, ub, color=color)
    plt.show()
    # for sub, fig in figmap.items():
    #     plt.figure(fig.number)
    #     plt.show()

def load_model(modelfile, objective):
    """Load model and set constraints."""
    model = cb.io.load_json_model(modelfile)
    model.objective = objective
    model.reactions.get_by_id(objective).upper_bound=1000

    # Set default exchange bounds from media composition
    init_cnc = dict()
    for rxn in model.reactions:
        if ( rxn.id.startswith('Ex_') and rxn.id.endswith('L') \
                or rxn.id in ['Ex_gly', 'Ex_his'] ):
            init_cnc[rxn.id] = rxn.upper_bound
            rxn.upper_bound *= 0.03
        if rxn.id in ['Ex_valL', 'Ex_ileL']:
            rxn.upper_bound=0
    model.reactions.Ex_glc.upper_bound=0
    model.reactions.Ex_cysL.upper_bound = 1000
    model.solver = 'glpk'
    return model

def dfba_main(params: MetaboliteCollection, model_file, objective_function, fba_method, substrates, 
              fva_run=False, tracked_reactions=[], tracked_metabolites=[], tmin_hours=0,
              tmax_hours=48, solutions_per_hour=1, dry_run=False):
    """Main function to compute dFBA solutions.
    Computes successive static FBA solutions and plots the estimated metabolite
    concentrations, uptake rates, and tracked reaction fluxes.

    Parameters:
    params -- dictionary mapping metabolite names to LogisticSet objects
            containing optimal logistic coefficients and errors
    tracked_rxns -- reactions to tracked and written to fluxes.xlsx
    fba_method -- FBA method to use for static solutions (function, default:
            cobra.flux_analysis.loopless.loopless_solution)
    fva_run -- whether to also compute an FVA solution
    modelfile -- location of metabolic model
    t_max -- end timepoint in hours (default 48)
    resolution -- number of static solutions per hour (default 1)
    obj -- reaction ID of objective function (default ATP_sink)
    dry_run -- True to avoid writing output to file (default False)
    """
    print(f"""dFBA log: Begin dFBA analysis with sheet, endpoint
          {tmax_hours} hours, and resolution {solutions_per_hour}.""")
    print(f"Using method {fba_method}.")
    if dry_run:
        print("Dry run, will not write results.")
    # Load metabolic model and logistic fit specs #
    print('dFBA log: loading model and specsheet...')
    model = load_model(parse_filepath(model_file), objective_function)
    
    # Set up logistic parameters and time scale
    nsol = int(round((tmax_hours - tmin_hours)*solutions_per_hour + 1, 0))
    ts_array = np.linspace(tmin_hours, tmax_hours, num=nsol)
    timecourse = list(ts_array)
    for mi, param_set in params.get_items():
        param_set.eval(ts_array)

    # Special-case adjustments for certain metabolites
    params.remove_met("5apn")
    if params.has_id("leuL") and params.has_id("valL") and params.has_id("isobuta"):
        for mi in "valL", "isobuta": # Remove these 4 lines if we get Val and Ile runs
            params.get_by_id(mi).tshift("Leucine", params.get_by_id("isocap").avg_x0("Leucine"))
    else:
        params.remove_met("valL")
        params.remove_met("isobuta")
    if params.has_id("proL") and params.has_id("ileL") and params.has_id("2mbut"):
        for mi in "ileL", "2mbut":
            params.get_by_id(mi).tshift("Leucine", params.get_by_id("proL").avg_x0("Proline"))
    else:
        params.remove_met("ileL")
        params.remove_met("2mbut")

    # Initialize data structure for tracked metabolites
    propdata = dict()
    propmets = []
    for mi in tracked_metabolites:
        if model.metabolites.has_id(mi + "_c"):
            propdata[mi] = {
                'rxns_in': set([]),
                'rxns_out': set([]),
                'data_in': [],
                'data_out': [],
            }
            propmets.append(model.metabolites.get_by_id(mi + "_c"))
            if fva_run:
                propdata[mi]['data_in_lb'] = []
                propdata[mi]['data_in_ub'] = []
                propdata[mi]['data_out_lb'] = []
                propdata[mi]['data_out_ub'] = []
        else:
            print(f"'{mi}_c' is not a valid metabolite.")

    # Initialize results dataframes
    mnames = [model.metabolites.get_by_id(mi + '_c').name for mi in params.get_ids()]
    results = [pd.DataFrame(0., index=timecourse, columns=mnames) for _ in range(6)]
    rnames = [model.reactions.get_by_id(ri).name for ri in tracked_reactions]
    rxnflux = pd.DataFrame(0., index=timecourse, columns=rnames)
    if fva_run:
        rxnf_ub = pd.DataFrame(0., index=timecourse, columns=rnames)
        rxnf_lb = pd.DataFrame(0., index=timecourse, columns=rnames)
        allrxns = [rxn.id for rxn in model.reactions]
        fullflux_lb = pd.DataFrame(0., index=timecourse, columns=allrxns)
        fullflux_ub = pd.DataFrame(0., index=timecourse, columns=allrxns)

    print('dFBA log: simulation output begin.')
    # Simulate static solutions over timecourse
    for i, t in enumerate(timecourse):
        # Update exchange flux bounds for each constrained metabolite
        for j, (mi, curveset) in enumerate(params.get_items()):
            ## Calculate and set exchange constraints ##
            fbounds = update_uptake_bounds(model, t, mi, curveset)
            mn = model.metabolites.get_by_id(mi + '_c').name

            ## Record exchange constraints ##
            if not (objective_function == "ATP_sink" and mi == 'glc'):
                for k, df in enumerate(results):
                    df.at[t, mn] = fbounds[k]

            ## Force butyrate pathways ##
            #  Ideally, this should be generalized by adding a "reaction" attribute to
            #  LogisticSet, so that a curve set associated with a particular metabolite
            #  and substrate can be used to also set bounds on reactions. This would be
            #  helpful for compounds produced by multiple substrates, where the
            #  contributing pathways are known, such as with butyrate.
            if mi == 'but':
                glc_but_l, glc_but_u, _, _ = curveset.get_bounds(t, substrate="Glucose")
                set_bounds(model, "ID_325", lower=glc_but_l, upper=glc_but_u, update=True)
                if "Threonine" in curveset.logistic_sets:
                    thr_but_l, thr_but_u, _, _ = curveset.get_bounds(t, substrate="Threonine")
                    set_bounds(model, "2HBD", lower=thr_but_l, upper=thr_but_u, update=True)

            ## Calculate glucose uptake constraint ##
            if objective_function == "ATP_sink" and mi in ['ac', 'eto', 'but', 'alaL']:
                signal, _, exch, _ = curveset.get_sol(t, substrate="Glucose")
                exch_l, exch_u, signal_l, signal_u = curveset.get_bounds(t, substrate="Glucose")
                gbounds = [exch, exch_l, exch_u, signal, signal_l, signal_u]
                for df, bound in zip(results, gbounds):
                    if mi == 'but':
                        df.at[t, 'beta-D-glucose'] += bound
                    else:
                        df.at[t, 'beta-D-glucose'] += 0.5*bound

        ## Set glucose uptake constraint ##
        if objective_function == "ATP_sink":
            set_bounds(model, "Ex_glc", lower=results[1].at[t, 'beta-D-glucose'],
                       upper=results[2].at[t, 'beta-D-glucose'], update=True)

        ## Get flux solution(s) and populate arrays ##
        sol = fba_method(model)
        if fva_run:
            sol_v = cb.flux_analysis.flux_variability_analysis(
                model,
                fraction_of_optimum=0.995,
                # loopless=False,
                loopless=True
            )
            fullflux_lb.loc[t, :] = sol_v['minimum']
            fullflux_ub.loc[t, :] = sol_v['maximum']
        if i % 10 == 0:
            print(f'dFBA log: Time = {t}  (cycle {i+1}) \tFBA solution: ' \
                  f'{sol.fluxes[objective_function]}')
        if sol.fluxes[objective_function] < 0.0001:
            print(f'dFBA log: infeasible solution on cycle {i}.')
        for ri in tracked_reactions:
            rn = model.reactions.get_by_id(ri).name
            rxnflux.at[t, rn] = sol.fluxes[ri]
            if fva_run:
                rxnf_ub.at[t, rn] = sol_v.at[ri, 'maximum']
                rxnf_lb.at[t, rn] = sol_v.at[ri, 'minimum']

        ## Display incremental solutions ##
        if t in [0, 6, 8, 12, 24, 36, 48]:
            print("time = " + str(i))
            print(model.summary(solution=sol))
            for met in propmets:
                print(met.summary(solution=sol))

        ## Record flux contributions for tracked metabolites ##
        for met in propmets:
            mi = met.id[:-2] # metabolite ID, without compartment label
            flux_dic = {}   # container for metabolite flux data
            # Collect contributions for metabolite influx and outflux
            for dr in ('in', 'out'):
                # fluxes in direction <dr>
                flux_dic[f"data_{dr}"] = dict()
                if fva_run:
                    flux_dic[f"data_{dr}_lb"] = dict()
                    flux_dic[f"data_{dr}_ub"] = dict()
            # Populate, considering all scenarios where met is produced/consumed
            for rxn in met.reactions:
                # Record outflux if meets threshold
                if (met in rxn.reactants and sol.fluxes[rxn.id] > 1E-5) \
                   or (met in rxn.products and sol.fluxes[rxn.id] < -1E-5):
                    flux_dic[f"data_out"][rxn.id] = rxn.metabolites[met] \
                                                   * sol.fluxes[rxn.id]
                    if fva_run:
                        flux_dic[f"data_out_lb"][rxn.id] = rxn.metabolites[met] \
                                                          * sol_v.at[rxn.id, 'minimum']
                        flux_dic[f"data_out_ub"][rxn.id] = rxn.metabolites[met] \
                                                          * sol_v.at[rxn.id, 'maximum']
                # Record influx if meets threshold
                elif (met in rxn.products and sol.fluxes[rxn.id] > 1E-5) \
                     or (met in rxn.reactants and sol.fluxes[rxn.id] < -1E-5):
                    flux_dic[f"data_in"][rxn.id] = rxn.metabolites[met] * \
                                                    sol.fluxes[rxn.id]
                    if fva_run:
                        flux_dic[f"data_in_lb"][rxn.id] = rxn.metabolites[met] \
                                                           * sol_v.at[rxn.id, 'minimum']
                        flux_dic[f"data_in_ub"][rxn.id] = rxn.metabolites[met] \
                                                           * sol_v.at[rxn.id, 'maximum']
            for dr in ('in', 'out'):
                propdata[mi][f'rxns_{dr}'].update(flux_dic[f"data_{dr}"].keys())
                propdata[mi][f'data_{dr}'].append(flux_dic[f'data_{dr}'])
                if fva_run:
                    propdata[mi][f'rxns_{dr}'].update(flux_dic[f"data_{dr}_lb"].keys())
                    propdata[mi][f'rxns_{dr}'].update(flux_dic[f"data_{dr}_ub"].keys())
                    propdata[mi][f'data_{dr}_lb'].append(flux_dic[f'data_{dr}_lb'])
                    propdata[mi][f'data_{dr}_ub'].append(flux_dic[f'data_{dr}_ub'])

        # FVA takes longer, print at end of each simulation
        if fva_run:
            now = datetime.datetime.now()
            print(f"t = {t:.2f} simulation complete ({now:%c})")

    ## Adjust flux solution directionality in output for visualization ##
    reverse_ids = ["ID_383", "ID_336", "ID_391", "HydEB"]
    reverse_nms = [model.reactions.get_by_id(rid).name for rid in reverse_ids]
    for rname in reverse_nms:
        if rname in rxnflux:
            rxnflux[rname] *= -1
            if fva_run:
                rxnf_ub[rname] *= -1
                rxnf_lb[rname] *= -1

    ## Plot run results ##
    print(f'Complete after {i} cycles ({t} hours). Final flux: '\
          f' {sol.fluxes["Ex_biomass"]}')
    areaplot(results[0], results[1], results[2], ylabel='flux (mM/h)')
    areaplot(results[3], results[4], results[5], ylabel='normalized signal')
    niceplot(rxnflux)

    ## Write run results ##
    if not dry_run:
        writer = pd.ExcelWriter(f'{SCDIR}/../../data/fluxes.xlsx', engine='openpyxl')
        snames = ('data', 'lbdev', 'ubdev', 'signal', 'slbdev', 'subdev')
        # Write estimated concentrations and fluxes
        for i, df in enumerate(results):
            df.to_excel(writer, sheet_name=snames[i], engine='openpyxl')
        # Write tracked reaction fluxes
        rxnflux.to_excel(writer, sheet_name='fluxes', engine='openpyxl')
        if fva_run:
            rxnf_ub.to_excel(writer, sheet_name='fluxub', engine='openpyxl')
            rxnf_lb.to_excel(writer, sheet_name='fluxlb', engine='openpyxl')
            # Write FVA bounds for ALL reactions to tab-delimited text
            fullflux_lb.to_csv(
                f'{SCDIR}/../../data/fullfluxlb.txt',
                sep='\t',
                index_label='Time (h)'
            )
            fullflux_ub.to_csv(
                f'{SCDIR}/../../data/fullfluxub.txt',
                sep='\t',
                index_label='Time (h)'
            )
        writer.close()
        # Open workbook to record flux fractions for tracked metabolites
        writer = pd.ExcelWriter(f'{SCDIR}/../../data/met_fluxes.xlsx', engine='openpyxl')
        wb=writer.book
        ws=wb.create_sheet("MetFluxes")
        writer.sheets["MetFluxes"] = ws
        startcol = 0

    # Record reaction flux fractions for tracked metabolites
    for met in propmets:
        mi = met.id[:-2]
        print(f"Flux slices for metabolite {met.name} ({met.id}).")
        flux_dfs = {}
        rnames = {}
        rids = {}
        names = ["Fluxes", "Flux Upper Bound", "Fluxes Lower Bound"]
        for dr in ('in', 'out'):
            rids[dr] = list(propdata[mi][f'rxns_{dr}'])
            rnames[dr] = [model.reactions.get_by_id(rid).name for rid in rids[dr]]
            labels = [f'data_{dr}']
            if fva_run:
                labels.extend([f'data_{dr}_lb', f'data_{dr}_ub'])
            for i, lab in enumerate(labels):
                # Convert flux fraction records to dataframes
                flux_dfs[lab] = pd.DataFrame.from_records(
                    propdata[mi][lab], 
                    index=timecourse
                ).fillna(0.)
                # Print result
                print(names[i])
                print(f"\t{dr}flux percentages:")
                v = flux_dfs[lab]
                totalflux = v.to_numpy().sum()
                cuts = v.sum(axis=0)
                for i, rxn in enumerate(rids[dr]):
                    print("\t", rxn, rnames[dr][i], f"{cuts.at[rxn]/totalflux*100:.2f}")
                print(f"\tAbsolute {dr}flux:")
                for i, rxn in enumerate(rids[dr]):
                    print("\t", rxn, rnames[dr][i], f"{cuts.at[rxn]}")
            # Write flux fraction results to Excel (met_fluxes.xlsx)
            if not dry_run:
                df = flux_dfs[f'data_{dr}']
                endcol = startcol + df.shape[1]
                if startcol == 0:
                    write_index = True
                    endcol += 1
                    ws.cell(row=1, column=2, value="Time")
                else:
                    write_index = False
                df.to_excel(writer, sheet_name="MetFluxes", startrow=2,
                            startcol=startcol, index=write_index)
                c = ws.cell(row=1, column=startcol+1, value=f"{met.name} {dr}flux")
                c.font = BOLD
                try:
                    ws.merge_cells(start_row=1, start_column=startcol+1, end_row=1,
                                   end_column=endcol)
                except ValueError:
                    pass
                colnames = list(df.columns)
                for i, j in enumerate(range(endcol-df.shape[1], endcol)):
                    rxname = model.reactions.get_by_id(colnames[i]).name
                    c = ws.cell(row=2, column=j+1, value=rxname)
                    c.font = BOLD
                startcol = endcol
    if not dry_run:
        writer.close()

def read_and_validate_cfg(cfg_filepath):
    parsed_path = parse_filepath(cfg_filepath)
    with open(parsed_path, "r") as rf:
        cfg = json.loads(rf.read())
    required_fields = ["method", "model", "objective_function", "nmr_substrates"]
    for field in required_fields:
        try:
            val = cfg[field]
        except KeyError:
            print(f"JSON config missing required field {field}.")
    return cfg

def parse_filepath(fpath):
    if not fpath.startswith('/'):
        return f"{SCDIR}/{fpath}"
    return fpath

def run_from_cfg(cfg_filepath):
    """Parse dFBA arguments from JSON config"""
    cfg = read_and_validate_cfg(cfg_filepath)
    params = MetaboliteCollection()
    substrates = [Substrate.from_json(s, params) for s in cfg.pop("nmr_substrates")]
    all_experiments = {}
    for substrate in substrates:
        all_experiments.update({parse_filepath(exp): substrate for exp in substrate.experiments})

    method = cfg.pop("method")
    if method == 'fva':
        fba_method = METHODS['fba']
        fva_run = True
    elif method in METHODS:
        fba_method = METHODS[method]
        fva_run = False
    else:
        print(f"Unsupported FBA method {method}. Please select one of "
                + ", ".join(METHODS))
        return
    sync_functions = synchronizers(
        [path for path in all_experiments], 
        stretch=cfg.pop("stretch", False), 
        plot=False
    )
    for tscale, (exp, substrate) in zip(sync_functions, all_experiments.items()):
        curves, errors = fit_trajectories(exp, substrate, tscale=tscale, plot=True)
        for met, pset in curves.items(): # update LogisticSet of metabolite
            params.get_by_name(met).add_curve(substrate.name, pset, errors[met])
    print("=========")
    print("dFBA run average curveshapes.")
    print("---------")
    for substrate in substrates:
        for met in substrate.metmap():
            params.get_by_name(met).set_runcount(substrate.name, len(substrate.experiments))
    for mid, curveset in params.get_items():
        print(mid)
        curveset.display_avg_coeffs()
    t_min = cfg["tmin_hours"]
    t_max = cfg["tmax_hours"]
    t_num = int(round((t_max-t_min)*cfg["solutions_per_hour"] + 1, 0))
    areaplot2(np.linspace(t_min, t_max, num=t_num), substrates, params)
    dfba_main(params, fva_run=fva_run, fba_method=fba_method, substrates=substrates, **cfg)

def handle_args():
    parser = argparse.ArgumentParser()
    parser.add_argument('cfgpath', metavar='CONFIG', help='path to dFBA config file')
    args = parser.parse_args()
    if not os.path.exists(args.cfgpath):
        print("Not a valid path " + args.cfgpath)
        return
    run_from_cfg(args.cfgpath)

if __name__ == "__main__":
    """Handle command line call"""
    handle_args()    
